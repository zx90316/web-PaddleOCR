#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PaddleOCR Web Interface
Copyright (c) 2025

This project provides a web interface for PaddleOCR.
This software uses PaddleOCR, which is licensed under Apache License 2.0.

PaddleOCR Copyright (c) 2020 PaddlePaddle Authors. All Rights Reserved.
Licensed under the Apache License, Version 2.0.
See: https://github.com/PaddlePaddle/PaddleOCR

本項目僅為 PaddleOCR 的網頁界面封裝，核心 OCR 功能由 PaddleOCR 提供。
"""

from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Request
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from typing import List, Optional
import json
import os
import tempfile
from paddlex import create_pipeline
import shutil
import numpy as np
from PIL import Image
import io
import uuid
from datetime import datetime
import database
import httpx  # 用於調用 CLIP 服務
import base64
import task_database as batch_db
import batch_processor
from urllib.parse import quote
import logging
from logging.handlers import RotatingFileHandler

# ==================== 日誌配置 ====================
# 建立 logs 目錄
os.makedirs("logs", exist_ok=True)

# 配置應用程式日誌
logger = logging.getLogger("paddleocr_app")
logger.setLevel(logging.INFO)

# 檔案處理器 - 使用輪替機制 (每個檔案 10MB, 保留 10 個備份)
file_handler = RotatingFileHandler(
    "logs/app.log",
    maxBytes=10*1024*1024,  # 10MB
    backupCount=10,
    encoding='utf-8'
)
file_handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(name)s - %(levelname)s - [%(filename)s:%(lineno)d] - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
))
logger.addHandler(file_handler)

# 主控台處理器
console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter(
    '%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
))
logger.addHandler(console_handler)

logger.info("=" * 60)
logger.info("PaddleOCR 應用程式日誌系統已初始化")
logger.info("=" * 60)

# ==================== FastAPI 應用程式初始化 ====================

# 初始化 FastAPI 應用程式
app = FastAPI(title="PaddleOCR 圖片識別服務", description="上傳圖片並提取指定的關鍵字")

# 設定靜態檔案服務
output_dir = "output"
os.makedirs(output_dir, exist_ok=True)

# 初始化資料庫
database.init_database()
batch_db.init_database()

app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/output", StaticFiles(directory=output_dir), name="output")

# 設定模板引擎
templates = Jinja2Templates(directory="templates")

# 初始化 PaddleOCR 管線
pipeline = create_pipeline(
    pipeline="./PP-ChatOCRv4-doc.yaml",
    initial_predictor=False
    )

# CLIP 服務配置
CLIP_SERVICE_URL = os.getenv("CLIP_SERVICE_URL", "http://192.168.80.24:8081")

# MLLM 服務配置
MLLM_SERVICE_URL = os.getenv("MLLM_SERVICE_URL", "http://localhost:8080")

async def check_mllm_health() -> bool:
    """
    檢查多模態大模型服務是否運行中
    Returns:
        bool: True 表示服務正常，False 表示服務不可用
    """
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            response = await client.get(f"{MLLM_SERVICE_URL}/health")
            if response.status_code == 200:
                data = response.json()
                # 檢查 errorCode 是否為 0 表示健康
                if data.get("errorCode") == 0:
                    logger.info(f"MLLM 服務健康檢查通過: {data.get('errorMsg', 'Healthy')}")
                    return True
                else:
                    logger.warning(f"MLLM 服務回應異常: errorCode={data.get('errorCode')}, errorMsg={data.get('errorMsg')}")
                    return False
            else:
                logger.warning(f"MLLM 服務健康檢查失敗: HTTP {response.status_code}")
                return False
    except httpx.TimeoutException:
        logger.error("MLLM 服務健康檢查超時")
        return False
    except Exception as e:
        logger.error(f"MLLM 服務健康檢查錯誤: {str(e)}")
        return False

# 請求模型
class OCRRequest(BaseModel):
    key_list: List[str]
    use_doc_orientation_classify: Optional[bool] = False
    use_doc_unwarping: Optional[bool] = False

# 響應模型
class OCRResponse(BaseModel):
    success: bool
    data: Optional[dict] = None
    error: Optional[str] = None

async def call_clip_service(pdf_file_path: str, positive_templates: List[UploadFile], negative_templates: List[UploadFile], positive_threshold: float, negative_threshold: float, skip_voided: bool = False, top_n_for_void_check: int = 5):
    """
    調用 CLIP 服務進行頁面匹配
    Args:
        pdf_file_path: PDF 文件路徑
        positive_templates: 正例範本圖片列表
        negative_templates: 反例範本圖片列表
        positive_threshold: 正例相似度閾值
        negative_threshold: 反例相似度閾值
        skip_voided: 是否跳過廢止頁面
        top_n_for_void_check: 檢查前 N 個候選頁面是否為廢止
    Returns:
        (matched_page_number, matched_page_image, matching_score, all_scores, voided_pages_checked)
    """
    async with httpx.AsyncClient(timeout=600.0, trust_env=False) as client:
        # 準備文件
        files = []

        # PDF 文件
        with open(pdf_file_path, 'rb') as f:
            files.append(('pdf_file', (os.path.basename(pdf_file_path), f.read(), 'application/pdf')))

        # 正例範本
        for template in positive_templates:
            await template.seek(0)  # 重置文件指針
            content = await template.read()
            files.append(('positive_templates', (template.filename, content, template.content_type)))

        # 反例範本
        for template in negative_templates:
            await template.seek(0)
            content = await template.read()
            files.append(('negative_templates', (template.filename, content, template.content_type)))

        # 準備表單數據
        data = {
            'positive_threshold': positive_threshold,
            'negative_threshold': negative_threshold,
            'skip_voided': skip_voided,
            'top_n_for_void_check': top_n_for_void_check
        }

        # 調用 CLIP 服務
        response = await client.post(
            f"{CLIP_SERVICE_URL}/match-page",
            files=files,
            data=data
        )

        if response.status_code != 200:
            raise HTTPException(status_code=500, detail=f"CLIP 服務調用失敗: {response.text}")

        result = response.json()

        if not result.get('success'):
            raise HTTPException(status_code=400, detail=result.get('error', '未知錯誤'))

        # 解碼 Base64 圖像
        matched_page_image = None
        if result.get('matched_page_base64'):
            img_data = base64.b64decode(result['matched_page_base64'])
            matched_page_image = Image.open(io.BytesIO(img_data))

        return (
            result.get('matched_page_number'),
            matched_page_image,
            result.get('matching_score'),
            result.get('all_page_scores', []),
            result.get('voided_pages_checked', [])
        )

async def perform_ocr_on_file(
    file_path: str,
    key_list_parsed: list,
    original_filename: str,
    task_output_dir: str,
    use_doc_orientation_classify: bool = False,
    use_doc_unwarping: bool = False,
    use_textline_orientation: bool = False,
    use_seal_recognition: bool = False,
    use_table_recognition: bool = True,
    use_llm: bool = True,
    use_mllm: bool = False,
):
    """
    對檔案執行 OCR 處理的核心邏輯
    Args:
        file_path: 圖片或PDF檔案路徑
        key_list_parsed: 已解析的關鍵字列表
        original_filename: 原始檔案名
        task_output_dir: 任務專屬的輸出目錄
        其他參數: OCR 處理選項
    Returns:
        處理結果字典
    """
    # 如果啟用 MLLM，先檢查服務是否可用
    if use_mllm and use_llm:
        logger.info("檢查 MLLM 服務健康狀態...")
        mllm_healthy = await check_mllm_health()
        if not mllm_healthy:
            logger.warning("MLLM 服務不可用，將退回使用標準 LLM")
            use_mllm = False
    # 執行視覺預測
    visual_predict_res = pipeline.visual_predict(
        input=file_path,
        use_doc_orientation_classify=use_doc_orientation_classify,
        use_doc_unwarping=use_doc_unwarping,
        use_textline_orientation=use_textline_orientation,
        use_common_ocr=True,
        use_seal_recognition=use_seal_recognition,
        use_table_recognition=use_table_recognition,
    )

    visual_info_list = []
    output_images = []
    for res in visual_predict_res:
        visual_info_list.append(res["visual_info"])
        layout_parsing_result = res["layout_parsing_result"]
        # 執行保存操作
        layout_parsing_result.save_to_img(task_output_dir)

        # 獲取保存後的檔案列表
        files = set(os.listdir(task_output_dir)) if os.path.exists(task_output_dir) else set()

        for file in files:
            if file.endswith('.png'):
                output_images.append(file)

    

    # 執行聊天查詢
    chat_result = {}
    if use_llm:
        if use_mllm:
            logger.info("使用 MLLM 進行多模態預測...")
            try:
                mllm_predict_res = pipeline.mllm_pred(
                    input=file_path,
                    key_list=key_list_parsed,
                )
                mllm_predict_info = mllm_predict_res["mllm_res"]
                logger.info("MLLM 預測完成，整合到聊天結果...")
            
                chat_result = pipeline.chat(
                    key_list=key_list_parsed,
                    visual_info=visual_info_list,
                    mllm_predict_info=mllm_predict_info,
                )
                logger.info("MLLM 整合聊天完成")
            except Exception as e:
                logger.error(f"MLLM 處理失敗: {str(e)}，退回標準 LLM")
                chat_result = pipeline.chat(
                    key_list=key_list_parsed,
                    visual_info=visual_info_list,
                )
        else:
            logger.info("使用標準 LLM 進行關鍵字提取...")
            chat_result = pipeline.chat(
                key_list=key_list_parsed,
                visual_info=visual_info_list,
            )
            logger.info("標準 LLM 提取完成")

    # 組合回應資料
    response_data = {
        "chat_result": chat_result.get("chat_res") if use_llm else None,
        "visual_info_list": visual_info_list,
        "key_list": key_list_parsed,
        "output_images": output_images,
        "original_filename": original_filename,
        "settings": {
            "use_doc_orientation_classify": use_doc_orientation_classify,
            "use_doc_unwarping": use_doc_unwarping,
            "use_textline_orientation": use_textline_orientation,
            "use_seal_recognition": use_seal_recognition,
            "use_table_recognition": use_table_recognition,
            "use_llm": use_llm,
            "use_mllm": use_mllm
        }
    }

    return response_data

@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    """返回上傳頁面"""
    return templates.TemplateResponse("index.html", {"request": request})

@app.post("/ocr", response_model=OCRResponse)
async def process_ocr(
    file: UploadFile = File(...),
    key_list: str = Form(...),
    use_doc_orientation_classify: bool = Form(False),
    use_doc_unwarping: bool = Form(False),
    use_textline_orientation: bool = Form(False),
    use_seal_recognition: bool = Form(False),
    use_table_recognition: bool = Form(True),
    use_llm: bool = Form(True),
    use_mllm: bool = Form(False),
):
    """處理圖片 OCR 請求"""
    temp_file_path = None
    task_id = str(uuid.uuid4())
    task_output_dir = os.path.join(output_dir, task_id)

    logger.info(f"收到 OCR 請求 - 任務ID: {task_id}, 檔案名稱: {file.filename}, 檔案類型: {file.content_type}")

    try:
        # 檢查檔案類型
        if not (file.content_type.startswith('image/') or file.content_type == 'application/pdf'):
            logger.warning(f"無效的檔案類型: {file.content_type} - 任務ID: {task_id}")
            raise HTTPException(status_code=400, detail="請上傳有效的圖片檔案或PDF檔案")

        # 解析關鍵字列表
        try:
            key_list_parsed = json.loads(key_list)
            logger.info(f"解析關鍵字列表成功 - 任務ID: {task_id}, 關鍵字數量: {len(key_list_parsed)}")
        except json.JSONDecodeError as e:
            logger.error(f"關鍵字列表解析失敗 - 任務ID: {task_id}, 錯誤: {str(e)}")
            raise HTTPException(status_code=400, detail="關鍵字列表格式錯誤")

        # 創建任務專屬輸出目錄
        os.makedirs(task_output_dir, exist_ok=True)
        logger.debug(f"創建輸出目錄 - 任務ID: {task_id}, 路徑: {task_output_dir}")

        # 創建臨時檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(file.filename)[1]) as temp_file:
            content = await file.read()
            temp_file.write(content)
            temp_file_path = temp_file.name

        logger.info(f"開始 OCR 處理 - 任務ID: {task_id}, 檔案大小: {len(content)} bytes")

        # 調用核心 OCR 處理函數
        response_data = await perform_ocr_on_file(
            file_path=temp_file_path,
            key_list_parsed=key_list_parsed,
            original_filename=file.filename,
            task_output_dir=task_output_dir,
            use_doc_orientation_classify=use_doc_orientation_classify,
            use_doc_unwarping=use_doc_unwarping,
            use_textline_orientation=use_textline_orientation,
            use_seal_recognition=use_seal_recognition,
            use_table_recognition=use_table_recognition,
            use_llm=use_llm,
            use_mllm=use_mllm
        )

        # 保存 response_data 到 JSON 檔案
        response_file = os.path.join(task_output_dir, "response.json")
        with open(response_file, 'w', encoding='utf-8') as f:
            json.dump(response_data, f, ensure_ascii=False, indent=2)

        # 添加 task_id 到回應
        response_data["task_id"] = task_id

        # 更新輸出圖片路徑為相對於 output 的路徑
        response_data["output_images"] = [f"{task_id}/{img}" for img in response_data["output_images"]]

        # 儲存任務資訊到資料庫
        database.insert_task(
            task_id=task_id,
            original_filename=file.filename,
            output_directory=task_output_dir,
            response_file=response_file,
            file_type='pdf' if file.content_type == 'application/pdf' else 'image',
            matched_page_number=None,
            settings=response_data["settings"]
        )

        logger.info(f"OCR 處理完成 - 任務ID: {task_id}, 檔案名稱: {file.filename}, 輸出圖片數量: {len(response_data['output_images'])}")
        return OCRResponse(success=True, data=response_data)

    except HTTPException as he:
        logger.warning(f"HTTP異常 - 任務ID: {task_id}, 狀態碼: {he.status_code}, 詳情: {he.detail}")
        raise
    except Exception as e:
        # 如果發生錯誤，清理輸出目錄
        logger.error(f"OCR 處理失敗 - 任務ID: {task_id}, 錯誤: {str(e)}", exc_info=True)
        if os.path.exists(task_output_dir):
            shutil.rmtree(task_output_dir)
        return OCRResponse(success=False, error=f"處理過程中發生錯誤: {str(e)}")
    finally:
        # 清理臨時檔案
        if temp_file_path and os.path.exists(temp_file_path):
            os.unlink(temp_file_path)
            logger.debug(f"清理臨時檔案 - 任務ID: {task_id}")

@app.post("/ocr-with-matching", response_model=OCRResponse)
async def process_ocr_with_matching(
    pdf_file: UploadFile = File(...),
    positive_templates: List[UploadFile] = File(...),
    negative_templates: List[UploadFile] = File(default=[]),
    key_list: str = Form(...),
    use_doc_orientation_classify: bool = Form(False),
    use_doc_unwarping: bool = Form(False),
    use_textline_orientation: bool = Form(False),
    use_seal_recognition: bool = Form(False),
    use_table_recognition: bool = Form(True),
    use_llm: bool = Form(True),
    positive_threshold: float = Form(0.25),
    negative_threshold: float = Form(0.30),
    skip_voided: bool = Form(False),
    top_n_for_void_check: int = Form(5),
    use_mllm: bool = Form(False),
):
    """
    處理 PDF 頁面匹配和 OCR 請求
    1. 接受 PDF 文件、正例範本圖片、反例範本圖片
    2. 調用 CLIP 服務找出最相似的頁面
    3. 對該頁面執行 OCR 處理
    """
    temp_pdf_path = None
    task_id = str(uuid.uuid4())
    task_output_dir = os.path.join(output_dir, task_id)

    try:
        # 檢查 PDF 檔案類型
        if pdf_file.content_type != 'application/pdf':
            raise HTTPException(status_code=400, detail="請上傳有效的 PDF 檔案")

        # 解析關鍵字列表
        try:
            key_list_parsed = json.loads(key_list)
        except json.JSONDecodeError:
            raise HTTPException(status_code=400, detail="關鍵字列表格式錯誤")

        # 創建任務專屬輸出目錄
        os.makedirs(task_output_dir, exist_ok=True)

        # 保存 PDF 到臨時檔案
        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_pdf:
            content = await pdf_file.read()
            temp_pdf.write(content)
            temp_pdf_path = temp_pdf.name

        # 調用 CLIP 服務進行頁面匹配
        print(f"調用 CLIP 服務進行頁面匹配...")
        best_page_number, best_page_image, best_score, all_scores, voided_pages = await call_clip_service(
            temp_pdf_path,
            positive_templates,
            negative_templates,
            positive_threshold,
            negative_threshold,
            skip_voided,
            top_n_for_void_check
        )

        if best_page_number is None:
            # 清理輸出目錄
            if os.path.exists(task_output_dir):
                shutil.rmtree(task_output_dir)
            return OCRResponse(
                success=False,
                error=f"未找到符合條件的頁面。請調整閾值參數。所有頁面分數: {all_scores}"
            )

        print(f"找到最佳匹配頁面: 第 {best_page_number} 頁, 分數: {best_score:.4f}")

        # 將匹配的頁面保存到任務輸出目錄
        matched_page_filename = f"matched_page_{best_page_number}.png"
        matched_page_path = os.path.join(task_output_dir, matched_page_filename)
        best_page_image.save(matched_page_path, 'PNG')

        # 調用核心 OCR 處理函數
        ocr_response_data = await perform_ocr_on_file(
            file_path=matched_page_path,
            key_list_parsed=key_list_parsed,
            original_filename=pdf_file.filename,
            task_output_dir=task_output_dir,
            use_doc_orientation_classify=use_doc_orientation_classify,
            use_doc_unwarping=use_doc_unwarping,
            use_textline_orientation=use_textline_orientation,
            use_seal_recognition=use_seal_recognition,
            use_table_recognition=use_table_recognition,
            use_llm=use_llm,
            use_mllm=use_mllm
        )

        # 在 OCR 結果中添加頁面匹配資訊
        response_data = {
            **ocr_response_data,  # 包含所有 OCR 結果
            "matched_page_number": best_page_number,
            "matching_score": float(best_score),
            "all_page_scores": all_scores,
            "matched_page_path": f"{task_id}/{matched_page_filename}",
        }

        # 如果有跳過的廢止頁面，添加到結果中
        if voided_pages:
            response_data["voided_pages_checked"] = voided_pages

        # 更新 settings 以包含匹配閾值
        response_data["settings"]["positive_threshold"] = positive_threshold
        response_data["settings"]["negative_threshold"] = negative_threshold
        response_data["settings"]["skip_voided"] = skip_voided
        response_data["settings"]["top_n_for_void_check"] = top_n_for_void_check

        # 保存 response_data 到 JSON 檔案
        response_file = os.path.join(task_output_dir, "response.json")
        with open(response_file, 'w', encoding='utf-8') as f:
            json.dump(response_data, f, ensure_ascii=False, indent=2)

        # 添加 task_id 到回應
        response_data["task_id"] = task_id

        # 更新輸出圖片路徑為相對於 output 的路徑
        response_data["output_images"] = [f"{task_id}/{img}" for img in response_data["output_images"]]

        # 儲存任務資訊到資料庫
        database.insert_task(
            task_id=task_id,
            original_filename=pdf_file.filename,
            output_directory=task_output_dir,
            response_file=response_file,
            file_type='pdf',
            matched_page_number=best_page_number,
            settings=response_data["settings"]
        )

        return OCRResponse(success=True, data=response_data)

    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_detail = f"處理過程中發生錯誤: {str(e)}\n{traceback.format_exc()}"
        print(error_detail)
        # 如果發生錯誤，清理輸出目錄
        if os.path.exists(task_output_dir):
            shutil.rmtree(task_output_dir)
        return OCRResponse(success=False, error=error_detail)

    finally:
        # 清理臨時檔案
        if temp_pdf_path and os.path.exists(temp_pdf_path):
            try:
                os.unlink(temp_pdf_path)
            except Exception as e:
                print(f"清理臨時檔案失敗: {temp_pdf_path}, 錯誤: {e}")

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    """返回管理後台頁面"""
    return templates.TemplateResponse("admin.html", {"request": request})

@app.get("/admin/tasks")
async def get_all_tasks(include_deleted: bool = False):
    """取得所有任務列表"""
    try:
        tasks = database.get_all_tasks(include_deleted=include_deleted)
        return {"success": True, "tasks": tasks}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/admin/task/{task_id}", response_class=HTMLResponse)
async def view_task_detail(request: Request, task_id: str):
    """查看任務詳情頁面"""
    task = database.get_task_by_id(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任務不存在")

    # 讀取 response.json
    response_data = None
    if os.path.exists(task['response_file']):
        with open(task['response_file'], 'r', encoding='utf-8') as f:
            response_data = json.load(f)

    return templates.TemplateResponse("task_detail.html", {
        "request": request,
        "task": task,
        "response_data": response_data
    })

@app.delete("/admin/task/{task_id}")
async def delete_task(task_id: str):
    """刪除任務"""
    try:
        # 取得任務資訊
        task = database.get_task_by_id(task_id)
        if not task:
            return {"success": False, "error": "任務不存在"}

        if task['is_deleted']:
            return {"success": False, "error": "任務已被刪除"}

        # 刪除實體檔案
        if os.path.exists(task['output_directory']):
            shutil.rmtree(task['output_directory'])

        # 標記資料庫為已刪除
        database.mark_task_deleted(task_id)

        return {"success": True, "message": "任務已刪除"}
    except Exception as e:
        return {"success": False, "error": str(e)}

# ==================== 批次任務管理 API ====================

@app.get("/batch-tasks", response_class=HTMLResponse)
async def batch_tasks_page(request: Request):
    """批次任務管理頁面"""
    return templates.TemplateResponse("batch_tasks.html", {"request": request})

@app.get("/batch-tasks/{task_id}/detail", response_class=HTMLResponse)
async def batch_task_detail_page(request: Request, task_id: str):
    """批次任務詳情頁面"""
    try:
        task = batch_db.get_task_by_id(task_id)
        if not task:
            raise HTTPException(status_code=404, detail="任務不存在")

        statistics = batch_db.get_task_statistics(task_id)
        keywords = batch_db.get_task_keywords(task_id)

        return templates.TemplateResponse("batch_task_detail.html", {
            "request": request,
            "task": task,
            "statistics": statistics,
            "keywords": keywords
        })
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/api/batch-tasks/create")
async def create_batch_task(
    task_name: str = Form(...),
    source_path: str = Form(...)
):
    """創建新的批次任務並掃描檔案"""
    logger.info(f"收到批次任務創建請求 - 任務名稱: {task_name}, 來源路徑: {source_path}")

    try:
        # 驗證路徑
        if not os.path.exists(source_path):
            logger.warning(f"批次任務創建失敗 - 路徑不存在: {source_path}")
            return {"success": False, "error": "指定的路徑不存在"}

        if not os.path.isdir(source_path):
            logger.warning(f"批次任務創建失敗 - 路徑不是目錄: {source_path}")
            return {"success": False, "error": "指定的路徑不是目錄"}

        # 創建任務
        task_id = str(uuid.uuid4())
        batch_db.create_batch_task(task_id, task_name, source_path)
        logger.info(f"批次任務已創建 - 任務ID: {task_id}, 任務名稱: {task_name}")

        # 掃描檔案
        files = batch_processor.scan_directory(source_path)
        logger.info(f"掃描目錄完成 - 任務ID: {task_id}, 找到 {len(files)} 個 PDF 檔案")

        if not files:
            logger.warning(f"批次任務創建失敗 - 未找到 PDF 檔案, 任務ID: {task_id}")
            return {"success": False, "error": "未找到任何 PDF 檔案"}

        # 添加檔案到任務
        batch_db.add_files_to_task(task_id, files)
        logger.info(f"批次任務創建成功 - 任務ID: {task_id}, 檔案數量: {len(files)}")

        return {
            "success": True,
            "task_id": task_id,
            "total_files": len(files),
            "message": f"成功創建任務，找到 {len(files)} 個 PDF 檔案"
        }

    except Exception as e:
        import traceback
        logger.error(f"批次任務創建失敗 - 任務名稱: {task_name}, 錯誤: {str(e)}", exc_info=True)
        return {"success": False, "error": f"創建任務失敗: {str(e)}\n{traceback.format_exc()}"}

@app.get("/api/batch-tasks")
async def get_batch_tasks(include_deleted: bool = False):
    """取得所有批次任務"""
    try:
        tasks = batch_db.get_all_tasks(include_deleted=include_deleted)
        return {"success": True, "tasks": tasks}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/api/batch-tasks/{task_id}")
async def get_batch_task_detail(task_id: str):
    """取得批次任務詳情"""
    try:
        task = batch_db.get_task_by_id(task_id)
        if not task:
            return {"success": False, "error": "任務不存在"}

        # 取得統計資訊
        stats = batch_db.get_task_statistics(task_id)

        # 取得關鍵字
        keywords = batch_db.get_task_keywords(task_id)

        return {
            "success": True,
            "task": task,
            "statistics": stats,
            "keywords": keywords
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/api/batch-tasks/{task_id}/keywords")
async def get_batch_task_keywords(task_id: str):
    """取得批次任務的關鍵字"""
    try:
        keywords = batch_db.get_task_keywords(task_id)
        return {"success": True, "keywords": keywords}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/api/batch-tasks/{task_id}/files")
async def get_batch_task_files(
    task_id: str,
    status: Optional[str] = None,
    stage1_status: Optional[str] = None,
    stage2_status: Optional[str] = None,
    limit: int = 50,
    offset: int = 0,
    exclude_base64: bool = True
):
    """
    取得任務的檔案列表

    Args:
        task_id: 任務ID
        status: 狀態篩選
        stage1_status: 第一階段狀態篩選
        stage2_status: 第二階段狀態篩選
        limit: 每頁數量(預設50,最大500)
        offset: 偏移量
        exclude_base64: 是否排除Base64圖片資料(預設True)
    """
    try:
        # 限制最大每頁數量,避免記憶體過載
        limit = min(limit, 500)

        # 取得檔案列表(不包含Base64以節省記憶體)
        files = batch_db.get_task_files(
            task_id,
            status=status,
            stage1_status=stage1_status,
            stage2_status=stage2_status,
            limit=limit,
            offset=offset,
            exclude_base64=exclude_base64
        )

        # 取得符合條件的總數量
        total_count = batch_db.get_task_files_count(
            task_id,
            status=status,
            stage1_status=stage1_status,
            stage2_status=stage2_status
        )

        return {
            "success": True,
            "files": files,
            "pagination": {
                "total": total_count,
                "limit": limit,
                "offset": offset,
                "has_more": (offset + limit) < total_count
            }
        }
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/batch-tasks/{task_id}/stage1/config")
async def configure_stage1(
    task_id: str,
    positive_templates: List[UploadFile] = File(...),
    negative_templates: List[UploadFile] = File(default=[]),
    positive_threshold: float = Form(0.25),
    negative_threshold: float = Form(0.30),
    skip_voided: bool = Form(False),
    top_n_for_void_check: int = Form(5)
):
    """配置第一階段參數"""
    try:
        # 將範本圖片轉換為 Base64
        positive_b64_list = []
        for template in positive_templates:
            content = await template.read()
            b64_str = base64.b64encode(content).decode('utf-8')
            positive_b64_list.append(b64_str)

        negative_b64_list = []
        for template in negative_templates:
            content = await template.read()
            b64_str = base64.b64encode(content).decode('utf-8')
            negative_b64_list.append(b64_str)

        # 保存配置
        config = {
            'positive_templates': positive_b64_list,
            'negative_templates': negative_b64_list,
            'positive_threshold': positive_threshold,
            'negative_threshold': negative_threshold,
            'skip_voided': skip_voided,
            'top_n_for_void_check': top_n_for_void_check
        }

        batch_db.save_task_stage1_config(task_id, config)

        return {"success": True, "message": "第一階段配置已保存"}

    except Exception as e:
        import traceback
        return {"success": False, "error": f"配置失敗: {str(e)}\n{traceback.format_exc()}"}

@app.post("/api/batch-tasks/{task_id}/stage2/config")
async def configure_stage2(
    task_id: str,
    keywords: str = Form(...),
    use_doc_orientation_classify: bool = Form(False),
    use_doc_unwarping: bool = Form(False),
    use_textline_orientation: bool = Form(False),
    use_seal_recognition: bool = Form(False),
    use_table_recognition: bool = Form(True),
    use_llm: bool = Form(True),
    use_mllm: bool = Form(False),
):
    """配置第二階段參數"""
    try:
        # 解析關鍵字
        keywords_list = json.loads(keywords)

        # 保存配置
        config = {
            'use_doc_orientation_classify': use_doc_orientation_classify,
            'use_doc_unwarping': use_doc_unwarping,
            'use_textline_orientation': use_textline_orientation,
            'use_seal_recognition': use_seal_recognition,
            'use_table_recognition': use_table_recognition,
            'use_llm': use_llm,
            'use_mllm': use_mllm
        }

        batch_db.save_task_stage2_config(task_id, config, keywords_list)

        return {"success": True, "message": "第二階段配置已保存"}

    except Exception as e:
        import traceback
        return {"success": False, "error": f"配置失敗: {str(e)}\n{traceback.format_exc()}"}

@app.post("/api/batch-tasks/{task_id}/stage1/start")
async def start_stage1_processing(task_id: str):
    """開始第一階段處理"""
    try:
        batch_processor.start_task_stage1(task_id, CLIP_SERVICE_URL)
        return {"success": True, "message": "第一階段處理已啟動"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/batch-tasks/{task_id}/stage2/start")
async def start_stage2_processing(task_id: str):
    """開始第二階段處理"""
    try:
        batch_processor.start_task_stage2(task_id)
        return {"success": True, "message": "第二階段處理已啟動"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/batch-tasks/{task_id}/pause")
async def pause_task_processing(task_id: str):
    """暫停任務"""
    try:
        batch_processor.pause_task(task_id)
        return {"success": True, "message": "任務已暫停"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/batch-tasks/{task_id}/stop")
async def stop_task_processing(task_id: str):
    """停止任務"""
    try:
        batch_processor.stop_task(task_id)
        return {"success": True, "message": "任務已停止"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/batch-tasks/{task_id}/stage1/restart")
async def restart_stage1_processing(task_id: str):
    """重新開始第一階段"""
    try:
        batch_processor.restart_task_stage1(task_id, CLIP_SERVICE_URL)
        return {"success": True, "message": "第一階段已重新啟動"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/batch-tasks/{task_id}/stage2/restart")
async def restart_stage2_processing(task_id: str):
    """重新開始第二階段"""
    try:
        batch_processor.restart_task_stage2(task_id)
        return {"success": True, "message": "第二階段已重新啟動"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.post("/api/batch-tasks/{task_id}/resume")
async def resume_failed_task(task_id: str):
    """
    從失敗點繼續執行任務

    此端點適用於任務中途失敗的情況,例如:
    - 執行到 1200/3000 時因錯誤中斷
    - 任務狀態變為 'failed'
    - 調用此端點後,從第 1201 個檔案繼續執行
    - 已完成和已失敗的檔案不會重新處理
    """
    try:
        batch_processor.resume_task_from_failure(task_id, CLIP_SERVICE_URL)

        # 取得任務資訊以返回詳細信息
        task = batch_db.get_task_by_id(task_id)
        stats = batch_db.get_task_statistics(task_id)

        stage = task['stage']
        if stage == 1:
            pending = stats.get('stage1_pending', 0)
            completed = stats.get('stage1_completed', 0)
            message = f"任務已從第一階段繼續執行。已完成: {completed}, 待處理: {pending}"
        else:
            pending = stats.get('stage2_pending', 0)
            completed = stats.get('stage2_completed', 0)
            message = f"任務已從第二階段繼續執行。已完成: {completed}, 待處理: {pending}"

        return {
            "success": True,
            "message": message,
            "stage": stage,
            "statistics": stats
        }
    except Exception as e:
        import traceback
        return {
            "success": False,
            "error": str(e),
            "traceback": traceback.format_exc()
        }

@app.delete("/api/batch-tasks/{task_id}")
async def delete_batch_task(task_id: str):
    """刪除批次任務"""
    try:
        batch_db.mark_task_deleted(task_id)
        return {"success": True, "message": "任務已刪除"}
    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/api/batch-tasks/{task_id}/files/{file_id}")
async def get_file_detail(task_id: str, file_id: int):
    """取得單個檔案的詳細資訊（包含圖片）"""
    try:
        conn = batch_db.get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT * FROM batch_files
            WHERE id = ? AND task_id = ?
        ''', (file_id, task_id))

        row = cursor.fetchone()
        if not row:
            return {"success": False, "error": "檔案不存在"}

        file_info = dict(row)

        # 解析 JSON 資料
        if file_info['ocr_result']:
            try:
                file_info['ocr_result'] = json.loads(file_info['ocr_result'])
            except (json.JSONDecodeError, TypeError, ValueError):  # nosec B110
                # 如果無法解析 JSON，保持原始字符串值
                pass

        if file_info['extracted_keywords']:
            try:
                file_info['extracted_keywords'] = json.loads(file_info['extracted_keywords'])
            except (json.JSONDecodeError, TypeError, ValueError):  # nosec B110
                # 如果無法解析 JSON，保持原始字符串值
                pass

        return {"success": True, "file": file_info}

    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/api/batch-tasks/{task_id}/files/{file_id}/image")
async def get_file_matched_image(task_id: str, file_id: int):
    """取得檔案的匹配頁面圖片"""
    try:
        from fastapi.responses import Response

        conn = batch_db.get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT matched_page_base64 FROM batch_files
            WHERE id = ? AND task_id = ?
        ''', (file_id, task_id))

        row = cursor.fetchone()
        if not row or not row['matched_page_base64']:
            return {"success": False, "error": "圖片不存在"}

        # 解碼 Base64
        img_data = base64.b64decode(row['matched_page_base64'])

        return Response(
            content=img_data,
            media_type="image/png",
            headers={"Content-Disposition": f"inline; filename=page_{file_id}.png"}
        )

    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/api/batch-tasks/{task_id}/files/{file_id}/pdf")
async def get_file_original_pdf(task_id: str, file_id: int):
    """取得檔案的原始 PDF"""
    try:
        from fastapi.responses import FileResponse
        import os

        conn = batch_db.get_connection()
        cursor = conn.cursor()

        cursor.execute('''
            SELECT file_path, file_name FROM batch_files
            WHERE id = ? AND task_id = ?
        ''', (file_id, task_id))

        row = cursor.fetchone()
        if not row or not row['file_path']:
            raise HTTPException(status_code=404, detail="檔案不存在")

        file_path = row['file_path']
        file_name = row['file_name']

        # 返回文件 - 使用URL編碼處理中文檔名
        
        encoded_filename = quote(file_name)

        headers = {
            'Content-Disposition': f'inline; filename*=UTF-8\'\'{encoded_filename}'
        }

        return FileResponse(
            path=file_path,
            media_type="application/pdf",
            headers=headers
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/batch-tasks/{task_id}/preview")
async def get_task_preview(task_id: str, limit: int = 10):
    """取得任務的預覽資訊（包含部分檔案的縮圖）"""
    try:
        # 取得已完成第一階段的檔案
        files = batch_db.get_task_files(
            task_id,
            stage1_status='completed',
            limit=limit
        )

        preview_data = []
        for f in files:
            preview_data.append({
                'id': f['id'],
                'file_name': f['file_name'],
                'matched_page_number': f['matched_page_number'],
                'matching_score': f['matching_score'],
                'stage2_status': f['stage2_status'],
                # Base64 圖片（可選擇性返回縮圖）
                'has_image': bool(f['matched_page_base64'])
            })

        return {"success": True, "files": preview_data}

    except Exception as e:
        return {"success": False, "error": str(e)}

@app.get("/api/batch-tasks/{task_id}/export")
async def export_task_to_excel(task_id: str):
    """匯出任務結果為 Excel (分批處理避免記憶體過載)"""
    try:
        from fastapi.responses import StreamingResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO

        # 取得任務資訊
        task = batch_db.get_task_by_id(task_id)
        if not task:
            return {"success": False, "error": "任務不存在"}

        # 取得關鍵字
        keywords = batch_db.get_task_keywords(task_id)

        # 創建 Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OCR 結果"

        # 設定標題樣式
        title_font = Font(bold=True, size=12)
        title_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        title_alignment = Alignment(horizontal="center", vertical="center")

        # 寫入標題行
        headers = ["檔案名稱", "檔案路徑", "狀態", "匹配頁面", "匹配分數"]
        headers.extend(keywords)
        headers.extend(["處理時間", "錯誤訊息"])

        # 初始化欄寬追蹤字典 (避免後續遍歷所有儲存格)
        column_widths = {}

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
            # 初始化欄寬為標題長度
            column_widths[col_idx] = len(str(header))

        # 分批處理檔案資料,避免一次載入過多記憶體
        batch_size = 100
        offset = 0
        row_idx = 2

        while True:
            # 每次只載入 100 筆資料,排除 Base64 圖片和 OCR 原始結果以提升效能
            files = batch_db.get_task_files(
                task_id,
                limit=batch_size,
                offset=offset,
                exclude_base64=True,
                exclude_ocr_result=True  # 匯出時不需要 OCR 原始結果,只需要提取的關鍵字
            )

            if not files:
                break

            # 寫入這批資料
            for file_info in files:
                # 欄位值列表
                values = [
                    file_info['file_name'],
                    file_info['file_path'],
                    file_info['status'],
                    file_info['matched_page_number'],
                    file_info['matching_score']
                ]

                # 解析提取的關鍵字
                extracted_keywords = {}
                if file_info['extracted_keywords']:
                    try:
                        extracted_keywords = json.loads(file_info['extracted_keywords'])
                    except (json.JSONDecodeError, TypeError, ValueError):  # nosec B110
                        # 如果無法解析 JSON，保持空字典
                        pass

                # 添加關鍵字值
                for keyword in keywords:
                    values.append(extracted_keywords.get(keyword, ""))

                # 添加處理時間和錯誤訊息
                values.append(file_info['processed_at'])
                values.append(file_info['error_message'])

                # 寫入儲存格並同步更新欄寬
                for col_idx, value in enumerate(values, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    # 同步追蹤最大欄寬
                    value_length = len(str(value)) if value is not None else 0
                    if value_length > column_widths.get(col_idx, 0):
                        column_widths[col_idx] = value_length

                row_idx += 1

            offset += batch_size

            # 如果這批資料少於 batch_size,表示已經是最後一批
            if len(files) < batch_size:
                break

        # 套用欄寬 (一次性設定,避免遍歷所有儲存格)
        from openpyxl.utils import get_column_letter
        for col_idx, max_length in column_widths.items():
            # 設定合理的欄寬範圍: 最小 10, 最大 50, 額外留 2 個字元空間
            adjusted_width = min(max(max_length + 2, 10), 50)
            column_letter = get_column_letter(col_idx)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 保存到內存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # 返回文件 - 使用URL編碼處理中文檔名
        filename = f"{task['task_name']}_{task_id[:8]}.xlsx"
        encoded_filename = quote(filename)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
        )

    except Exception as e:
        import traceback
        return {"success": False, "error": f"匯出失敗: {str(e)}\n{traceback.format_exc()}"}

@app.get("/health")
async def health_check():
    """健康檢查端點"""
    return {"status": "healthy", "message": "PaddleOCR 服務運行正常"}

@app.get("/health/logs")
async def log_health_check():
    """日誌健康檢查端點"""
    try:
        from log_monitor import LogMonitor

        monitor = LogMonitor()
        result = monitor.check_log_health()
        summary = monitor.get_status_summary()

        return {
            "success": True,
            "health_check": result,
            "summary": summary
        }
    except Exception as e:
        logger.error(f"日誌健康檢查失敗: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/admin/cleanup")
async def trigger_data_cleanup(dry_run: bool = False):
    """觸發資料清理作業"""
    try:
        from data_retention import DataCleanupManager

        logger.info(f"收到資料清理請求 (dry_run={dry_run})")
        manager = DataCleanupManager(dry_run=dry_run)
        result = manager.cleanup_all()

        logger.info(f"資料清理完成: 刪除 {result['summary']['total_deleted']} 項, 釋放 {result['summary']['total_freed_mb']:.2f} MB")

        return {
            "success": True,
            "result": result
        }
    except Exception as e:
        logger.error(f"資料清理失敗: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e)
        }

@app.get("/admin/retention-policy")
async def get_retention_policy():
    """取得資料保存政策"""
    try:
        from data_retention import RetentionPolicy

        policy = RetentionPolicy()
        return {
            "success": True,
            "policies": policy.policies
        }
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

@app.post("/admin/retention-policy")
async def update_retention_policy(policies: dict):
    """更新資料保存政策"""
    try:
        from data_retention import RetentionPolicy

        policy = RetentionPolicy()
        for category, days in policies.items():
            policy.set_retention_days(category, int(days))
        policy.save_policies()

        logger.info(f"資料保存政策已更新: {policies}")

        return {
            "success": True,
            "message": "保存政策已更新"
        }
    except Exception as e:
        logger.error(f"更新保存政策失敗: {e}", exc_info=True)
        return {
            "success": False,
            "error": str(e)
        }

if __name__ == "__main__":
    import uvicorn
    import os

    # 從環境變量讀取配置，默認只綁定 localhost
    # 生產環境若需要對外訪問，請設置環境變量 APP_HOST=0.0.0.0
    host = os.getenv("APP_HOST", "127.0.0.1")
    port = int(os.getenv("APP_PORT", "8080"))

    print("🚀 啟動 PaddleOCR 網站服務...")
    print(f"🌐 服務地址: http://{host}:{port}")
    print(f"🌐 本機訪問: http://localhost:{port}")
    print(f"🛠️ 管理後台: http://localhost:{port}/admin")

    # nosec B104: 從環境變量讀取 host，默認為安全的 127.0.0.1
    # 只有明確設置環境變量才會綁定到所有接口，並會顯示警告
    if host == "0.0.0.0":  # nosec B104
        print("⚠️  警告: 服務綁定到所有網絡接口 (0.0.0.0)，請確保已設置適當的防火牆規則")

    uvicorn.run(app, host=host, port=port)